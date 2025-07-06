#!/usr/bin/python

import re
import docx2txt
import os
import openpyxl

BasePath = r"C:\Users\dboliveira\OneDrive - BRASS DO BRASIL\Projetos\1 - Em Desenvolvimento\BdB220500\Consolidacao\BdB200301"

workbook = openpyxl.Workbook()

regex = re.compile(
    (
        r"(\d\d\d-\D\D\D-\D\D\d\d\.\d\d/d\d.\d\d/d\d.\d\d /d\d\.\d\d/d\d\.\d\d /d\d\.\d\d /d\d\.\d\d /d\d\d\.\d\d /d\d\d-\D\D\D-\D\D\d\d\.\d\d)|"
        r"(\d\d\d-\D\D-\d\d\s\/\s\d\d\s\/\s\d\d)|"
        r"(\d\d\d-\D\D\D\D\D\d\d\.\d\d\/\d\d)|" # 439-PSVBB46.01/02
        r"(\d\d\d-\D\D\D\D\D\d\d\.\d\d\/\d\d)|"
        r"(\d\d\d-\s\D\D-\D\D\d\d\.\d\d@\d\d)|"
        r"(\d\d\d-\s\D\D\D\D-\D\D\d\d\.\d\d)|"
        r"(\d\d\d-\D\D\D-\D\D\d\d\d\D\.\d\d)|"
        r"(\d\d\d-\D\D\D \D\D\d\d\D\.\d\d)|"
        r"(\d\d\d-\s\D\D\D-\D\D\d\d\.\d\d)|"
        r"(\d\d\d-\D\D\D-\D\D\d\d\D\.\d\d)|"
        r"(\d\d\d-\D\D\D-\D\D\D\d\d\.\d\d)|"
        r"(\d\d\d-\D\D-\D\D\d\d\d\D\.\d\d)|"
        r"(\d\d\d-\D\D\D-\D\D\d\d\d\.\d\d)|"
        r"(\d\d\d-\D\D\D\D\D\d\d\.\/\d\d)|"
        r"(\d\d\d-\D\D\D\D\D\d\d\.\d\d)|"
        r"(\d\d\d-\D\D\D-\D\D\d\d\.\d\d)|"
        r"(\d\d\d-\D\D\DD\D\d\d\D\.\d\d)|"
        r"(\d\d\d-\D\D-\D\D\d\d\d\.\d\d)|"
        r"(\d\d\d-\D\D\D-\d\d\d\.\d\d)|"
        r"(\d\d\d-\D\D-\d\d\d\D\/\D)|"
        r"(\d\d\d-\D\D-\d\d\d\D)|"
        r"(\d\d\d-\D\D-\d\d\d)|"
        r"(\d\d\d-\D\D-\d\d)"
    )
    , flags = re.I
)
Files = {
    "TXT": [
        "BdB200301-0117-V-CP0001.txt",
        "BdB200301-0117-V-EG0001.txt",
        "BdB200301-0117-V-EG0002.txt",
        "BdB200301-0117-V-EG0003.txt",
        "BdB200301-0117-V-EG0005.txt",
        "BdB200301-0117-V-EG0006.txt",
        "BdB200301-0117-V-EG0007.txt",
        "BdB200301-0117-V-ET0001.txt",
        "BdB200301-0117-V-ET0002.txt",
        "BdB200301-0117-V-ET0003.txt",
        "BdB200301-0117-V-ET0004.txt",
        "BdB200301-0117-V-MC0001.txt",
        "BdB200301-0117-V-MC0002.txt",
        "BdB200301-0117-V-MC0003.txt",
        "BdB200301-0117-V-MC0004.txt",
        "BdB200301-0117-V-MC0005.txt",
        "BdB200301-0117-V-MC0006.txt",
        "BdB200301-0117-V-MC0007.txt",
        "BdB200301-0117-V-MC0008.txt",
        "BdB200301-0117-V-MC0009.txt",
        "BdB200301-0117-V-MC0010.txt",
        "BdB200301-0120-V-EG0001.txt",
        "BdB200301-0120-V-EG0002.txt",
        "BdB200301-0120-V-EG0003.txt",
        "BdB200301-0120-V-EG0005.txt",
        "BdB200301-0120-V-EG0006.txt",
        "BdB200301-0120-V-EG0007.txt",
        "BdB200301-0120-V-ET0001.txt",
        "BdB200301-0120-V-MC0001.txt",
        "BdB200301-0120-V-MC0002.txt",
        "BdB200301-0120-V-MC0003.txt",
        "BdB200301-0120-V-MC0004.txt",
        "BdB200301-0120-V-MC0006.txt",
        "BdB200301-0120-V-MC0007.txt",
        "BdB200301-0120-V-MC0008.txt",
        "BdB200301-0120-V-MC0011.txt",
        "BdB200301-0120-V-MC0012.txt",
        "BdB200301-0120-V-MC0014.txt",
        "BdB200301-0120-V-MC0015.txt",
        "BdB200301-0120-V-MC0016.txt",
        "BdB200301-0120-V-MC0017.txt",
    ],
    "XLS": [
        "BdB200301-0117-V-FD0001.xlsx",
        "BdB200301-0117-V-FD0002.xlsx",
        "BdB200301-0117-V-FD0003.xlsx",
        "BdB200301-0117-V-FD0004.xlsx",
        "BdB200301-0117-V-FD0006.xlsx",
        "BdB200301-0117-V-FD0007.xlsx",
        "BdB200301-0117-V-FD0008.xlsx",
        "BdB200301-0117-V-FD0009.xlsx",
        "BdB200301-0117-V-FD0010.xlsx",
        "BdB200301-0117-V-FD0011.xlsx",
        "BdB200301-0117-V-FD0012.xlsx",
        "BdB200301-0117-V-FD0013.xlsx",
        "BdB200301-0117-V-FD0014.xlsx",
        "BdB200301-0117-V-FD0015.xlsx",
        "BdB200301-0117-V-FD0016.xlsx",
        "BdB200301-0117-V-FD0017.xlsx",
        "BdB200301-0117-V-FD0018.xlsx",
        "BdB200301-0117-V-FD0019.xlsx",
        "BdB200301-0117-V-FD0020.xlsx",
        "BdB200301-0117-V-FE0002.xlsx",
        "BdB200301-0117-V-FE0003.xlsx",
        "BdB200301-0117-V-FE0004.xlsx",
        "BdB200301-0117-V-FE0005.xlsx",
        "BdB200301-0117-V-FE0006.xlsx",
        "BdB200301-0117-V-FE0007.xlsx",
        "BdB200301-0117-V-FE0011.xlsx",
        "BdB200301-0117-V-FE0012.xlsx",
        "BdB200301-0117-V-FP0001.xlsx",
        "BdB200302-0117-V-LL0001.xlsx",
        "BdB200301-0120-V-FD0001.xlsx",
        "BdB200301-0120-V-FD0002.xlsx",
        "BdB200301-0120-V-FD0003.xlsx",
        "BdB200301-0120-V-FD0004.xlsx",
        "BdB200301-0120-V-FD0005.xlsx",
        "BdB200301-0120-V-FD0006.xlsx",
        "BdB200301-0120-V-FD0007.xlsx",
        "BdB200301-0120-V-FD0008.xlsx",
        "BdB200301-0120-V-FD0009.xlsx",
        "BdB200301-0120-V-FD0010.xlsx",
        "BdB200301-0120-V-FD0011.xlsx",
        "BdB200301-0120-V-FD0012.xlsx",
        "BdB200301-0120-V-FD0014.xlsx",
        "BdB200301-0120-V-FD0017.xlsx",
        "BdB200301-0120-V-FD0018.xlsx",
        "BdB200301-0120-V-FD0019.xlsx",
        "BdB200301-0120-V-FE0002.xlsx",
        "BdB200301-0120-V-FE0003.xlsx",
        "BdB200301-0120-V-FE0004.xlsx",
        "BdB200301-0120-V-FE0005.xlsx",
        "BdB200301-0120-V-FE0006.xlsx",
        "BdB200301-0120-V-FE0007.xlsx",
        "BdB200301-0120-V-FE0008.xlsx",
        "BdB200301-0120-V-FE0009.xlsx",
        "BdB200301-0120-V-FE0010.xlsx",
        "BdB200301-0120-V-FE0011.xlsx",
        "BdB200301-0120-V-FE0012.xlsx",
        "BdB200301-0120-V-FE0013.xlsx",
        "BdB200301-0120-V-FE0014.xlsx",
        "BdB200301-0120-V-FE0015.xlsx",
        "BdB200301-0120-V-FE0016.xlsx",
        "BdB200301-0120-V-FE0019.xlsx",
        "BdB200301-0120-V-FE0020.xlsx",
        "BdB200301-0120-V-FP0001.xlsx",
    ],
}

worksheet = workbook.active
worksheet.cell(row = 1, column = 1).value = "TAG"
worksheet.cell(row = 1, column = 2).value = "Documento"
worksheet.cell(row = 1, column = 3).value = "Observações"

Counter = 2
# TXT
for file in Files["TXT"]:
    try:
        if os.path.splitext(file) == "docx" or os.path.splitext(file) == "doc":
            text = docx2txt.process(BasePath+"\\"+file)
        else:
            text = open(BasePath+"\\"+file, "r", encoding = 'latin-1')
            text = text.read()
        
        for match in regex.finditer(text):
            if match:
                workbook.active = workbook.sheetnames.index("Sheet")
                worksheet = workbook.active
                worksheet.cell(row = Counter, column = 1).value = match.group()
                worksheet.cell(row = Counter, column = 2).value = file
                Counter += 1
    except Exception as error:
        print(f"There was a problem in processing the file {file}.")
        #print(error)
        pass
    else:
        print(f"The file {file} have been processed sucessfully.")

# XLSX
for file in Files["XLS"]:
    try:
        workbook02 = openpyxl.load_workbook(BasePath+"\\"+file)
        
        for sheet in workbook02.worksheets:
            for row in range(0, sheet.max_row):
                for col in sheet.iter_cols(1, sheet.max_column):
                    for match in regex.finditer(str(col[row].value)):
                        if match:
                            workbook.active = workbook.sheetnames.index("Sheet")
                            worksheet = workbook.active
                            worksheet.cell(row = Counter, column = 1).value = match.group()
                            worksheet.cell(row = Counter, column = 2).value = file
                            worksheet.cell(row = Counter, column = 3).value = f"Planilha: {sheet.title}, Celula: {col[0].coordinate}"
                            Counter += 1
        
        workbook02.close()
    except Exception as error:
        print(f"There was a problem in processing the file {file}.")
        #print(error)
        pass
    else:
        print(f"The file {file} have been processed sucessfully.")

workbook.save(BasePath+"\\"+"BdB200301 - TAGs.xlsx")
workbook.close()





