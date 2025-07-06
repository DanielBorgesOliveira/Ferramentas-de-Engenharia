#!/usr/bin/python

#pip install openpyxl pandas

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

# Recursive function to create nested dictionaries from multiple levels of '->' splits
def add_to_nested_dict(current_dict, categories, data):
    """ Recursively adds categories to the nested dictionary structure """
    if len(categories) != 1:
        category = categories[0]
        if category not in current_dict:
            current_dict[category] = {}
        add_to_nested_dict(current_dict[category], categories[1:], data)
    else:
        current_dict[categories[0]] = data

# Function to read the worksheet into a generalized structured format
def read_worksheet_to_structure(file_path, equipment_name):
    # Read the Excel file into a pandas DataFrame
    df = pd.read_excel(file_path, header=0)
    
    # Check if DataFrame is empty
    #print("DataFrame shape:", df.shape)  # Debugging: Check the DataFrame shape
    
    structured_data = []  # List to store structured data dictionaries
    # Dictionary to store the current row's data
    current_dict = {}
    current_layout = {}
    
    # Loop through the rows to structure the data
    for index, row in df.iterrows():
        if equipment_name in row.get('Equipamento') or "Geral" in row.get('Equipamento'):
            description = row.get('Descrição')  # Safely get the 'Descrição' column
            if pd.notna(description):
                categories = [part.strip() for part in description.split('->')]  # Split by '->' into categories
                
                if categories[0] not in current_layout:
                    current_layout[categories[0]] = [row['Layout'], row['Equipamento']]
                
                # Add this row's data into the nested dictionary structure
                add_to_nested_dict(current_dict, categories, [row['Un.'], row['Especificado']])
                #row.get('style', '')
    
    # If there is any data, append the last dictionary
    if current_dict:
        return (current_dict, current_layout)
    else:
        return None

# Function to process the data and write to a structured Excel file with formatting
def write_excel_vertitical_layout(wb, structured_data, structured_layout, equipment):
    # Counter for the item column
    item_counter = 1
    
    # Horizontal layout column counter
    horizontal_layout_counter = 1
    
    # Vertical layout row counter
    vertical_layout_counter = 1
    
    # Recursive function to write subcategories with Excel's indentation and gray background for the first level
    def vertical_layout_write_subcategories(ws, subcategories, level = 0):
        """
        Recursive function to write subcategories with Excel's indentation and gray background for the first level
        """
        nonlocal item_counter
        nonlocal vertical_layout_counter
        
        for subcategory, values in subcategories.items():
            # Recursively write categories items.
            if isinstance(values, dict):
                # Write the data
                #ws.cell(row = vertical_layout_counter, column = 1, value = item_counter)
                ws.cell(row = vertical_layout_counter, column = 1).border = thin_border
                ws.cell(row = vertical_layout_counter, column = 1).alignment = Alignment(horizontal = "center", vertical = "center")
                ws.cell(row = vertical_layout_counter, column = 1).font = Font(bold=True)
                
                ws.cell(row = vertical_layout_counter, column = 2, value = subcategory)
                ws.cell(row = vertical_layout_counter, column = 2).border = thin_border
                ws.cell(row = vertical_layout_counter, column = 2).alignment = Alignment(horizontal = "left", vertical = "center", indent = level * 4, wrap_text = True)
                ws.cell(row = vertical_layout_counter, column = 2).font = Font(bold=True)
                
                #ws.cell(row = vertical_layout_counter, column = 3, value = "")
                ws.cell(row = vertical_layout_counter, column = 3).border = thin_border
                ws.cell(row = vertical_layout_counter, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                ws.cell(row = vertical_layout_counter, column = 3).font = Font(bold=True)
                
                #ws.cell(row = vertical_layout_counter, column = 4, value = "")
                ws.cell(row = vertical_layout_counter, column = 4).border = thin_border
                ws.cell(row = vertical_layout_counter, column = 4).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                ws.cell(row = vertical_layout_counter, column = 4).font = Font(bold=True)
                
                vertical_layout_counter += 1
                
                vertical_layout_write_subcategories(ws, values, level + 1)
            
            # Write the items and stop the recursive call.
            elif isinstance(values, list):
                # Write the data
                ws.cell(row = vertical_layout_counter, column = 1, value = item_counter)
                ws.cell(row = vertical_layout_counter, column = 1).border = thin_border
                ws.cell(row = vertical_layout_counter, column = 1).alignment = Alignment(horizontal = "center", vertical = "center")
                
                ws.cell(row = vertical_layout_counter, column = 2, value = subcategory)
                ws.cell(row = vertical_layout_counter, column = 2).border = thin_border
                ws.cell(row = vertical_layout_counter, column = 2).alignment = Alignment(horizontal = "left", vertical = "center", indent = level * 4, wrap_text = True)
                
                ws.cell(row = vertical_layout_counter, column = 3, value = values[0])
                ws.cell(row = vertical_layout_counter, column = 3).border = thin_border
                ws.cell(row = vertical_layout_counter, column = 3).alignment = Alignment(horizontal = "center", vertical = "center")
                
                ws.cell(row = vertical_layout_counter, column = 4, value = values[1])
                ws.cell(row = vertical_layout_counter, column = 4).border = thin_border
                ws.cell(row = vertical_layout_counter, column = 4).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                
                vertical_layout_counter += 1
                
                item_counter += 1
    
    # Recursive function to write subcategories with Excel's indentation and gray background for the first level
    def horizontal_layout_write_subcategories(ws, subcategories, level = 0):
        """
        There is one level for subcategories in horizontal layout.
        """
        nonlocal item_counter
        nonlocal horizontal_layout_counter
        
        for subcategory, values in subcategories.items():
            if isinstance(values, list):
                # Write the data
                ws.cell(row = 2, column = horizontal_layout_counter, value = item_counter)
                ws.cell(row = 2, column = horizontal_layout_counter).border = thin_border
                ws.cell(row = 2, column = horizontal_layout_counter).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                ws.cell(row = 2, column = horizontal_layout_counter).fill = gray_fill
                #ws.cell(row = 2, column = horizontal_layout_counter).font = Font(bold=True)
                
                ws.cell(row = 3, column = horizontal_layout_counter, value = subcategory)
                ws.cell(row = 3, column = horizontal_layout_counter).border = thin_border
                ws.cell(row = 3, column = horizontal_layout_counter).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                ws.cell(row = 3, column = horizontal_layout_counter).fill = gray_fill
                #ws.cell(row = 3, column = horizontal_layout_counter).font = Font(bold=True)
                
                ws.cell(row = 4, column = horizontal_layout_counter, value = values[0])
                ws.cell(row = 4, column = horizontal_layout_counter).border = thin_border
                ws.cell(row = 4, column = horizontal_layout_counter).alignment = Alignment(horizontal = "center", vertical = "center")
                ws.cell(row = 4, column = horizontal_layout_counter).fill = gray_fill
                #ws.cell(row = 4, column = horizontal_layout_counter).font = Font(bold=True)
                
                ws.cell(row = 5, column = horizontal_layout_counter, value = values[1])
                ws.cell(row = 5, column = horizontal_layout_counter).border = thin_border
                ws.cell(row = 5, column = horizontal_layout_counter).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                #ws.cell(row = 5, column = horizontal_layout_counter).fill = gray_fill
                #ws.cell(row = 5, column = horizontal_layout_counter).font = Font(bold=True)
                
                horizontal_layout_counter += 1
                
                item_counter += 1  # Increment the item counter
    
    def set_columns_width(ws, size = 15):
        if type(size) == int:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column letter
                ws.column_dimensions[column].width = size  # Set the column width
        elif type(size) == tuple:
            ws.column_dimensions["A"].width = size[0]
            ws.column_dimensions["B"].width = size[1]
            ws.column_dimensions["C"].width = size[2]
            ws.column_dimensions["D"].width = size[3]
    
    # Iterate through the structured data and write it into the worksheet
    for category, subcategories in structured_data.items():
        if structured_layout[category][0] == "Vertical":
            # Creates the worksheet if it not exists
            if f"{equipment}-V" not in wb.sheetnames:
                ws_vertical = wb.create_sheet(title = f"{equipment}-V")
                
                # Write headers
                ws_vertical.cell(row = vertical_layout_counter, column = 1, value="Item")
                ws_vertical.cell(row = vertical_layout_counter, column = 1).border = thin_border
                ws_vertical.cell(row = vertical_layout_counter, column = 1).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                ws_vertical.cell(row = vertical_layout_counter, column = 1).fill = gray_fill
                ws_vertical.cell(row = vertical_layout_counter, column = 1).font = Font(bold=True)
                
                ws_vertical.cell(row = vertical_layout_counter, column = 2, value="Descrição")
                ws_vertical.cell(row = vertical_layout_counter, column = 2).border = thin_border
                ws_vertical.cell(row = vertical_layout_counter, column = 2).alignment = Alignment(horizontal = "left", vertical = "center", wrap_text = True)
                ws_vertical.cell(row = vertical_layout_counter, column = 2).fill = gray_fill
                ws_vertical.cell(row = vertical_layout_counter, column = 2).font = Font(bold=True)
                
                ws_vertical.cell(row = vertical_layout_counter, column = 3, value="Un.")
                ws_vertical.cell(row = vertical_layout_counter, column = 3).border = thin_border
                ws_vertical.cell(row = vertical_layout_counter, column = 3).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                ws_vertical.cell(row = vertical_layout_counter, column = 3).fill = gray_fill
                ws_vertical.cell(row = vertical_layout_counter, column = 3).font = Font(bold=True)
                
                ws_vertical.cell(row = vertical_layout_counter, column = 4, value="Especificado")
                ws_vertical.cell(row = vertical_layout_counter, column = 4).border = thin_border
                ws_vertical.cell(row = vertical_layout_counter, column = 4).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                ws_vertical.cell(row = vertical_layout_counter, column = 4).fill = gray_fill
                ws_vertical.cell(row = vertical_layout_counter, column = 4).font = Font(bold=True)
                
                vertical_layout_counter += 1
            
            # Write the main category and apply formatting
            #ws_vertical.cell(row = vertical_layout_counter, column = 1, value = item_counter)
            ws_vertical.cell(row = vertical_layout_counter, column = 1).border = thin_border
            ws_vertical.cell(row = vertical_layout_counter, column = 1).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
            ws_vertical.cell(row = vertical_layout_counter, column = 1).fill = gray_fill
            ws_vertical.cell(row = vertical_layout_counter, column = 1).font = Font(bold=True)
                
            ws_vertical.cell(row = vertical_layout_counter, column = 2, value = category)
            ws_vertical.cell(row = vertical_layout_counter, column = 2).border = thin_border
            ws_vertical.cell(row = vertical_layout_counter, column = 2).alignment = Alignment(horizontal = "left", vertical = "center", wrap_text = True)
            ws_vertical.cell(row = vertical_layout_counter, column = 2).fill = gray_fill
            ws_vertical.cell(row = vertical_layout_counter, column = 2).font = Font(bold=True)
            
            ws_vertical.cell(row = vertical_layout_counter, column = 3, value="")
            ws_vertical.cell(row = vertical_layout_counter, column = 3).border = thin_border
            ws_vertical.cell(row = vertical_layout_counter, column = 3).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
            ws_vertical.cell(row = vertical_layout_counter, column = 3).fill = gray_fill
            ws_vertical.cell(row = vertical_layout_counter, column = 3).font = Font(bold=True)
            
            ws_vertical.cell(row = vertical_layout_counter, column = 4, value="")
            ws_vertical.cell(row = vertical_layout_counter, column = 4).border = thin_border
            ws_vertical.cell(row = vertical_layout_counter, column = 4).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
            ws_vertical.cell(row = vertical_layout_counter, column = 4).fill = gray_fill
            ws_vertical.cell(row = vertical_layout_counter, column = 4).font = Font(bold=True)
            vertical_layout_counter += 1
            
            # Write the subcategories recursively
            vertical_layout_write_subcategories(ws_vertical, subcategories)
            
            set_columns_width(ws_vertical, size = (10, 50, 10, 20))
        
        elif structured_layout[category][0] == "Horizontal":
            # Creates the worksheet if it not exists
            if f"{equipment}-H" not in wb.sheetnames:
                ws_horizontal = wb.create_sheet(title = f"{equipment}-H")
                
                # Write headers
                ws_horizontal.cell(row = 1, column = horizontal_layout_counter, value="Título")
                ws_horizontal.cell(row = 1, column = horizontal_layout_counter).border = thin_border
                ws_horizontal.cell(row = 1, column = horizontal_layout_counter).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                ws_horizontal.cell(row = 1, column = horizontal_layout_counter).fill = gray_fill
                ws_horizontal.cell(row = 1, column = horizontal_layout_counter).font = Font(bold=True)
                
                ws_horizontal.cell(row = 2, column = horizontal_layout_counter, value="Item")
                ws_horizontal.cell(row = 2, column = horizontal_layout_counter).border = thin_border
                ws_horizontal.cell(row = 2, column = horizontal_layout_counter).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                ws_horizontal.cell(row = 2, column = horizontal_layout_counter).fill = gray_fill
                ws_horizontal.cell(row = 2, column = horizontal_layout_counter).font = Font(bold=True)
                
                ws_horizontal.cell(row = 3, column = horizontal_layout_counter, value="Descrição")
                ws_horizontal.cell(row = 3, column = horizontal_layout_counter).border = thin_border
                ws_horizontal.cell(row = 3, column = horizontal_layout_counter).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                ws_horizontal.cell(row = 3, column = horizontal_layout_counter).fill = gray_fill
                ws_horizontal.cell(row = 3, column = horizontal_layout_counter).font = Font(bold=True)
                
                ws_horizontal.cell(row = 4, column = horizontal_layout_counter, value="Un.")
                ws_horizontal.cell(row = 4, column = horizontal_layout_counter).border = thin_border
                ws_horizontal.cell(row = 4, column = horizontal_layout_counter).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                ws_horizontal.cell(row = 4, column = horizontal_layout_counter).fill = gray_fill
                ws_horizontal.cell(row = 4, column = horizontal_layout_counter).font = Font(bold=True)
                
                ws_horizontal.cell(row = 5, column = horizontal_layout_counter, value="Especificado")
                ws_horizontal.cell(row = 5, column = horizontal_layout_counter).border = thin_border
                ws_horizontal.cell(row = 5, column = horizontal_layout_counter).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
                ws_horizontal.cell(row = 5, column = horizontal_layout_counter).font = Font(bold=True)
                
                horizontal_layout_counter += 1
            
            # Write the main category and apply formatting
            ws_horizontal.cell(row = 1, column = horizontal_layout_counter, value = category)
            ws_horizontal.cell(row = 1, column = horizontal_layout_counter).border = thin_border
            ws_horizontal.cell(row = 1, column = horizontal_layout_counter).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
            ws_horizontal.cell(row = 1, column = horizontal_layout_counter).font = Font(bold=True)
            ws_horizontal.cell(row = 1, column = horizontal_layout_counter).fill = gray_fill
            
            horizontal_layout_write_subcategories(ws_horizontal, subcategories)
            
            set_columns_width(ws_horizontal, size = 15)

#equipments = [
#    "Compressor Parafuso",
#    "Vaso de Pressao",
#    "Valvula de Alivio",
#    "Motor",
#]
equipments = [
    "Filtro de Areia",
    "Válvula Guilhotina",
]

file_path = r'C:\Users\dboliveira\OneDrive - BRASS DO BRASIL\1-EXECUCAO DE PROJETO\Documentos Padronizados\FD\FD - Geral.xlsx'
output_file = r'C:\Users\dboliveira\Desktop\1\Structured_Output_Final.xlsx'

# Step 1: Read the worksheet into a structured format
structured_data = []
structured_layout = []

for equipment_name in equipments:
    data, layout = read_worksheet_to_structure(file_path, equipment_name)
    structured_data.append(data)
    structured_layout.append(layout)

# Step 2: Process the structured data and write it to the formatted Excel output
wb = Workbook() # Create a new Excel workbook
for i in range(len(structured_data)):
    write_excel_vertitical_layout(wb, structured_data[i], structured_layout[i], equipments[i])

wb.remove(wb.worksheets[0])
wb.save(output_file) # Save the workbook to the specified output file


